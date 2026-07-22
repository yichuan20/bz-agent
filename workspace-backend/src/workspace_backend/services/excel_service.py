"""Excel service — load, patch, and save .xlsx workbooks.

The ``_xlsx_to_api`` fast-parse and ``_sidecar_to_api`` functions are copied verbatim
from the old app.py.  The ``ExcelService`` wrapper is the only new code.

Cell format on the wire (CellData, matches ExcelViewSheetArea):
    { value?, formula?, fontBold?, fontItalic?, fontColor?, bgColor?, align?,
      dataFormatString?, wrapText?, fontSize? }

Sidecar format (internal, from excel-worker / our writes):
    { v?, f?, s?: { bold?, italic?, fg?, bg?, align?, format?, wrap?, fontSize? } }

``_sidecar_to_api`` converts sidecar → CellData.  ``_xlsx_to_api`` reads an .xlsx
directly and returns CellData (for files that have no sidecar yet).
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any

# ── Constants ─────────────────────────────────────────────────────────────────

_SIDECAR_SUFFIX = ".excel.json"

# ── Formula recalc (delegates to excel-worker.py engine, loaded once) ────────


def _recalc_sheets(sheets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Re-evaluate all formula cells across all sheets. Best-effort; returns original on error."""
    try:
        import importlib.util

        _scripts = Path(__file__).resolve().parents[4] / "bzcode_assets" / "scripts"
        _worker_path = _scripts / "excel-worker.py"
        _spec = importlib.util.spec_from_file_location("_excel_worker", str(_worker_path))
        if _spec and _spec.loader:
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)  # type: ignore[union-attr]
            updated, _ = _mod.recalc_all_sheets(sheets)
            return updated
    except Exception:
        pass
    return sheets


def _sidecar_path(p: Path) -> Path:
    return p.parent / f".{p.name}{_SIDECAR_SUFFIX}"


# ── sidecar → API (copied verbatim from old app.py _sidecar_to_api) ──────────


def _sidecar_to_api(sidecar: dict, path: Path | None = None) -> dict:
    """Convert sidecar JSON → API schema the frontend (ExcelViewSheetArea) expects."""
    sheets = []
    for sheet in sidecar.get("sheets", []):
        api_cells: dict = {}
        for ref, cd in sheet.get("cells", {}).items():
            api_cd: dict = {}
            v = cd.get("v")
            if v is not None:
                api_cd["value"] = v
            if "f" in cd:
                api_cd["formula"] = cd["f"]
            s = cd.get("s", {})
            if s.get("bold"):
                api_cd["fontBold"] = True
            if s.get("italic"):
                api_cd["fontItalic"] = True
            if s.get("fg"):
                _fg = s["fg"]
                if _fg.startswith("#") and len(_fg) == 7:
                    _fg = "FF" + _fg[1:]
                api_cd["fontColor"] = _fg
            if s.get("bg"):
                _bg = s["bg"]
                if _bg.startswith("#") and len(_bg) == 7:
                    _bg = "FF" + _bg[1:]
                api_cd["bgColor"] = _bg
            if s.get("align"):
                api_cd["align"] = s["align"]
            if s.get("format") is not None:
                api_cd["dataFormatString"] = s["format"]
            if s.get("wrap") is not None:
                api_cd["wrapText"] = bool(s["wrap"])
            if s.get("fontSize") is not None:
                api_cd["fontSize"] = s["fontSize"]
            if api_cd:
                api_cells[ref] = api_cd

        # Grid dimensions: prefer explicit grid field, fall back to col_widths array
        grid_obj = sheet.get("grid", {})
        col_widths: dict = {str(k): v for k, v in (grid_obj.get("columnIndexToWidth") or {}).items()}
        if not col_widths:
            # Old-style col_widths (list) or new-style columnIndexToWidth (dict)
            raw_cw = sheet.get("columnIndexToWidth") or sheet.get("col_widths") or {}
            if isinstance(raw_cw, dict):
                for k, w in raw_cw.items():
                    if w:
                        col_widths[str(k)] = max(30, int(w * 7.5) if w < 50 else int(w))
            elif isinstance(raw_cw, list):
                for i, w in enumerate(raw_cw):
                    if w:
                        col_widths[str(i)] = max(30, int(w * 7.5) if w < 50 else int(w))
        row_heights: dict = {
            str(k): v for k, v in (grid_obj.get("rowIndexToHeight") or sheet.get("rowIndexToHeight") or {}).items()
        }

        # Merged cell ranges
        merged = sheet.get("mergedCells") or sheet.get("mergedCellRanges") or []

        stem = path.stem if path else sidecar.get("name", "")
        sheets.append(
            {
                "sheetName": sheet.get("sheetName") or sheet.get("name", ""),
                "cells": api_cells,
                "images": [],
                "columnIndexToWidth": col_widths,
                "rowIndexToHeight": row_heights,
                "hiddenColIndices": [],
                "hiddenRowIndices": [],
                "mergedCellIndices": [],
                "mergedCellRanges": merged,
            }
        )
    stem = path.stem if path else sidecar.get("name", "")
    return {"id": stem, "name": stem, "sheets": sheets, "sources": []}


# ── Direct xlsx parse → API format (from old app.py excel_load slow-path) ─────


def _xlsx_to_api(p: Path) -> dict:
    """Parse an .xlsx file directly → API CellData format (no sidecar needed)."""
    import openpyxl  # type: ignore[import-untyped]

    wb_vals = openpyxl.load_workbook(str(p), data_only=True)
    wb_forms = openpyxl.load_workbook(str(p), data_only=False)
    formula_map: dict = {}
    for ws_f in wb_forms.worksheets:
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_map[(ws_f.title, cell.coordinate)] = cell.value
    sheets = []
    for ws in wb_vals.worksheets:
        cells: dict = {}
        col_widths: dict = {}
        for row in ws.iter_rows():
            for cell in row:
                cell_id = cell.coordinate
                formula = formula_map.get((ws.title, cell_id))
                cached_val = cell.value
                if cached_val is None and not formula:
                    continue
                cd: dict = {}
                if cached_val is not None:
                    cd["value"] = str(cached_val) if not isinstance(cached_val, (int, float, bool)) else cached_val
                if formula:
                    cd["formula"] = formula
                try:
                    f = cell.font
                    if f.bold:
                        cd["fontBold"] = True
                    if f.italic:
                        cd["fontItalic"] = True
                    if f.color and f.color.type == "rgb":
                        rgb = f.color.rgb  # AARRGGBB
                        if rgb[:2].upper() not in ("00", ""):
                            cd["fontColor"] = rgb
                except Exception:
                    pass
                try:
                    fill = cell.fill
                    if fill and fill.fill_type == "solid" and fill.fgColor and fill.fgColor.type == "rgb":
                        rgb = fill.fgColor.rgb
                        if rgb not in ("FF000000", "00000000", "FFFFFFFF"):
                            cd["bgColor"] = rgb
                except Exception:
                    pass
                if cd:
                    cells[cell_id] = cd
        for col_letter, dim in (ws.column_dimensions or {}).items():
            if dim.width:
                idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                col_widths[str(idx)] = max(30, int(dim.width * 7.5))
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        sheets.append(
            {
                "sheetName": ws.title,
                "cells": cells,
                "images": [],
                "columnIndexToWidth": col_widths,
                "rowIndexToHeight": {},
                "hiddenColIndices": [],
                "hiddenRowIndices": [],
                "mergedCellIndices": [],
                "mergedCellRanges": merged_ranges,
            }
        )
    wb_vals.close()
    wb_forms.close()
    return {"id": p.stem, "name": p.stem, "sheets": sheets, "sources": []}


# ── Sidecar write helpers ─────────────────────────────────────────────────────


def _read_sidecar(p: Path) -> dict[str, Any] | None:
    sc = _sidecar_path(p)
    if not sc.exists():
        return None
    try:
        return json.loads(sc.read_text(encoding="utf-8"))
    except OSError, json.JSONDecodeError:
        return None


def _write_sidecar(p: Path, data: dict[str, Any]) -> None:
    sc = _sidecar_path(p)
    sc.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def _write_xlsx(p: Path, sidecar: dict[str, Any]) -> None:
    """Regenerate .xlsx from sidecar (best-effort)."""
    import openpyxl  # type: ignore[import-untyped]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in sidecar.get("sheets", []):
        ws = wb.create_sheet(title=str(sheet.get("sheetName", "Sheet")))
        for ref, cell in (sheet.get("cells") or {}).items():
            v = cell.get("v")
            f = cell.get("f")
            if f:
                ws[ref] = f"={f}"
            elif v is not None:
                ws[ref] = v
        for rng in sheet.get("mergedCellRanges") or []:
            try:
                ws.merge_cells(str(rng))
            except Exception:
                pass
    wb.save(str(p))
    wb.close()


# ── ExcelService — thin wrapper ───────────────────────────────────────────────


class ExcelService:
    async def load(self, path: str) -> dict[str, Any]:
        p = Path(path)

        def _do() -> dict[str, Any]:
            sc = _read_sidecar(p)
            if sc is not None:
                return _sidecar_to_api(sc, p)
            result = _xlsx_to_api(p)
            return result

        return await asyncio.to_thread(_do)

    async def patch(self, path: str, sheet: str | None, cells: dict[str, Any]) -> dict[str, Any]:
        p = Path(path)

        def _do() -> dict[str, Any]:
            sc = _read_sidecar(p)
            if sc is None:
                # Build a minimal sidecar from direct xlsx parse — convert CellData back to {v,f,s}
                api = _xlsx_to_api(p)
                sc = {"sheets": []}
                for sh in api["sheets"]:
                    sidecar_cells: dict = {}
                    for ref, cd in sh["cells"].items():
                        entry: dict = {}
                        if "value" in cd:
                            entry["v"] = cd["value"]
                        if "formula" in cd:
                            entry["f"] = cd["formula"]
                        if entry:
                            sidecar_cells[ref] = entry
                    sc["sheets"].append(
                        {
                            "sheetName": sh["sheetName"],
                            "cells": sidecar_cells,
                            "mergedCellRanges": sh.get("mergedCellRanges", []),
                            "columnIndexToWidth": sh.get("columnIndexToWidth", {}),
                            "rowIndexToHeight": sh.get("rowIndexToHeight", {}),
                        }
                    )

            sheets = sc.get("sheets", [])
            target = next((s for s in sheets if s.get("sheetName") == sheet), sheets[0] if sheets else None)
            if target is None:
                target = {
                    "sheetName": sheet or "Sheet1",
                    "cells": {},
                    "columnIndexToWidth": {},
                    "rowIndexToHeight": {},
                    "mergedCellRanges": [],
                }
                sheets.append(target)
            existing_cells = target.setdefault("cells", {})
            for ref, data in cells.items():
                if data is None:
                    existing_cells.pop(ref, None)
                else:
                    existing_cells[ref] = {**existing_cells.get(ref, {}), **data}
            sc["sheets"] = sheets

            # Re-evaluate all formula cells using the same engine as excel-worker.py
            # (imported directly — no subprocess needed).
            sc["sheets"] = _recalc_sheets(sc["sheets"])

            _write_sidecar(p, sc)
            _write_xlsx(p, sc)
            return _sidecar_to_api(sc, p)

        return await asyncio.to_thread(_do)

    async def set_grid(
        self, path: str, sheet: str | None, col_widths: dict[str, Any], row_heights: dict[str, Any]
    ) -> None:
        p = Path(path)

        def _do() -> None:
            sc = _read_sidecar(p) or {"sheets": []}
            sheets = sc.get("sheets", [])
            target = next((s for s in sheets if s.get("sheetName") == sheet), sheets[0] if sheets else None)
            if target is not None:
                if col_widths:
                    target["columnIndexToWidth"] = col_widths
                if row_heights:
                    target["rowIndexToHeight"] = row_heights
            _write_sidecar(p, sc)

        await asyncio.to_thread(_do)

    async def merge_cells(self, path: str, sheet: str | None, merged_cells: list[str]) -> dict[str, Any]:
        p = Path(path)

        def _do() -> dict[str, Any]:
            sc = _read_sidecar(p) or {"sheets": []}
            sheets = sc.get("sheets", [])
            target = next((s for s in sheets if s.get("sheetName") == sheet), sheets[0] if sheets else None)
            if target is not None:
                target["mergedCellRanges"] = merged_cells
            _write_sidecar(p, sc)
            _write_xlsx(p, sc)
            return _sidecar_to_api(sc, p)

        return await asyncio.to_thread(_do)

    async def rename_sheet(self, path: str, old_name: str, new_name: str) -> str:
        p = Path(path)

        def _do() -> str:
            sc = _read_sidecar(p) or {"sheets": []}
            sheets = sc.get("sheets", [])
            names = [s.get("sheetName") for s in sheets]
            if new_name in names:
                raise ValueError(f"Sheet '{new_name}' already exists")
            for s in sheets:
                if s.get("sheetName") == old_name:
                    s["sheetName"] = new_name
            _write_sidecar(p, sc)
            return new_name

        return await asyncio.to_thread(_do)

    async def add_sheet(self, path: str, sheet_name: str | None = None) -> str:
        p = Path(path)

        def _do() -> str:
            sc = _read_sidecar(p) or {"sheets": []}
            sheets = sc.setdefault("sheets", [])
            existing = {s.get("sheetName") for s in sheets}
            name = sheet_name or "Sheet2"
            n = 2
            while name in existing:
                name = f"Sheet{n}"
                n += 1
            sheets.append(
                {
                    "sheetName": name,
                    "cells": {},
                    "columnIndexToWidth": {},
                    "rowIndexToHeight": {},
                    "mergedCellRanges": [],
                }
            )
            _write_sidecar(p, sc)
            return name

        return await asyncio.to_thread(_do)
