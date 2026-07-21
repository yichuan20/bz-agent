#!/usr/bin/env python3
from __future__ import annotations

"""
create-xlsx.py — generate an Excel workbook from JSON data.

The agent provides data as JSON, then calls this script to produce a properly
formatted XLSX without burning tokens on formatting details.

Usage:
  # Data from a JSON file
  python3 bzcode/scripts/create-xlsx.py --data-file /tmp/data.json --out /path/to/output.xlsx

  # Inline JSON for small datasets
  python3 bzcode/scripts/create-xlsx.py --data '{"sheets":[...]}' --out output.xlsx

Data JSON schema:
  {
    "sheets": [
      {
        "name": "Sheet1",
        "headers": ["Col A", "Col B", "Col C"],
        "rows": [
          ["val1", "val2", "val3"],
          ["val4", "val5", "val6"]
        ],
        "formulas": {                   # optional: cell -> formula string
          "C2": "=A2+B2",
          "C3": "=SUM(A2:A10)"
        },
        "col_widths": [20, 15, 10],     # optional: per-column width in characters
        "styles": {                     # optional: per-cell style overrides
          "A2": { "bg": "#FFF3E0", "fg": "#E65100" },
          "B3": { "bold": true, "align": "center" }
        }
      }
    ]
  }

Output:
  Prints JSON: { "ok": true, "path": "/abs/path/to/output.xlsx", "sheets": N, "rows": N }
"""
import argparse
import json
import sys
from pathlib import Path


def make_workbook(data: dict, out_path: Path) -> tuple[int, int]:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1473DF", bgColor="1473DF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER_SIDE = Side(style="thin", color="D0D0D0")
    CELL_BORDER = Border(
        left=BORDER_SIDE,
        right=BORDER_SIDE,
        top=BORDER_SIDE,
        bottom=BORDER_SIDE,
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet
    wb.calculation.calcMode = "auto"
    wb.calculation.calcOnSave = True
    wb.calculation.fullCalcOnLoad = True

    total_rows = 0
    for sheet_def in data.get("sheets", []):
        ws = wb.create_sheet(title=sheet_def.get("name", "Sheet"))
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        formulas = sheet_def.get("formulas", {})
        col_widths = sheet_def.get("col_widths", [])

        # Write header row
        if headers:
            for c_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = CELL_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = CELL_BORDER
            total_rows += 1

        # Apply formulas (overwrite cell values)
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula

        # Column widths
        for c_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        # Auto-fit columns not explicitly set (based on header length)
        if not col_widths:
            for c_idx, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = max(12, len(str(header)) + 4)

        # Freeze the header row
        if headers:
            ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return len(data.get("sheets", [])), total_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate an XLSX workbook from JSON data.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--data", help="Inline JSON data string.")
    group.add_argument("--data-file", help="Path to a JSON data file.")
    parser.add_argument("--out", required=True, help="Output .xlsx path.")
    args = parser.parse_args()

    out_path = Path(args.out)
    if out_path.suffix.lower() not in (".xlsx", ".xls"):
        out_path = out_path.with_suffix(".xlsx")

    # Prefer excel-worker.py which creates both the xlsx and the sidecar JSON
    # (required for the frontend to display computed cross-sheet formula values).
    worker = Path(__file__).parent / "excel-worker.py"
    if worker.exists():
        import subprocess

        cmd = [sys.executable, str(worker)]
        if args.data_file:
            cmd += ["--data-file", args.data_file]
        else:
            cmd += ["--data", args.data]
        cmd += ["--out", str(out_path)]
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            try:
                w = json.loads(result.stdout)
                print(
                    json.dumps(
                        {
                            "ok": True,
                            "path": w.get("xlsx_path", str(out_path.resolve())),
                            "sheets": w.get("sheets", 0),
                            "rows": 0,
                        }
                    )
                )
                return
            except Exception:
                pass
        # If worker failed, fall through to openpyxl fallback
        if result.stderr:
            sys.stderr.write(result.stderr)

    if args.data_file:
        f = Path(args.data_file)
        if not f.exists():
            print(json.dumps({"error": f"data file not found: {f}"}))
            sys.exit(1)
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            print(json.dumps({"error": f"invalid JSON: {e}"}))
            sys.exit(1)
    else:
        try:
            data = json.loads(args.data)
        except json.JSONDecodeError as e:
            print(json.dumps({"error": f"invalid JSON: {e}"}))
            sys.exit(1)

    try:
        sheets, rows = make_workbook(data, out_path)
    except Exception as exc:
        print(json.dumps({"error": f"could not create workbook: {exc}"}))
        sys.exit(1)

    print(json.dumps({"ok": True, "path": str(out_path.resolve()), "sheets": sheets, "rows": rows}))


if __name__ == "__main__":
    main()
